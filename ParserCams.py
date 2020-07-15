from selenium import webdriver
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.keys import Keys
import time
from bs4 import BeautifulSoup
import requests
import openpyxl
import os

chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument('headless')
chrome_options.binary_location = os.environ.get('GOOGLE_CHROME_BIN')
chrome_options.add_argument('--headless')
chrome_options.add_argument('--disable-dev-shm-usage')
chrome_options.add_argument('--no-sandbox')
driver = webdriver.Chrome(executable_path=os.environ.get('CHROMEDRIVER_PATH'), chrome_options=chrome_options)


def kenex_parser():
    kenex_results = []
    i = 1
    while i <= 35:
        url = 'https://www.kenex.cz/kamery/strana-{number}'+'/'
        url = url.format(number=i)
        req = requests.get(url)
        html = req.content
        soup = BeautifulSoup(html, 'lxml')
        cams = soup.find_all('li', class_='product')

        for x in cams:
            name = x.find('a', class_='p-name').find('span').get_text(strip=True).lower()
            availability = x.find('span', class_='p-cat-availability').get_text(strip=True).lower()
            if 'obj' in availability:
                availability = 'Не на складе'
            else:
                availability = 'На складе'
            href = x.find('a', class_='p-name').get('href').lower()
            kenex_results.append({'name': name, 'availability': availability, 'href': 'https://www.kenex.cz'+href})

        i += 1

    for i in kenex_results:
        i['name'] = i['name'].replace(',', '.')
    return kenex_results


def viakom_parser(url, number):
    viakom_results = []
    driver.get(url)
    elem = driver.find_element_by_tag_name('body')
    number_of_scrolls = number

    while number_of_scrolls:
        elem.send_keys(Keys.PAGE_DOWN)
        time.sleep(0.9)
        number_of_scrolls -= 1

    html = driver.page_source
    soup = BeautifulSoup(html, 'lxml')
    cams = soup.find_all('article', class_='product-item')

    for x in cams:
        try:
            name = x.find('h2', class_='product-name').find('a').get_text(strip=True).lower()

            try:
                try:
                    availability = x.find('span', class_='stock').get_text(strip=True).lower()
                except:
                    availability = x.find('span', class_='nostock').get_text(strip=True).lower()
            except:
                availability = x.find('div', class_='product-stock').get_text(strip=True).lower()

            if 'skladem' in availability:
                availability = 'На складе'
            else:
                availability = 'Не на складе'

            href = x.find('h2', class_='product-name').find('a').get('href').lower()
            viakom_results.append({'name': name, 'availability': availability, 'href': 'https://www.viakom.cz'+href})
        except:
            pass

    for i in viakom_results:
        i['name'] = i['name'].replace(',', '.')
    return viakom_results


def eurosat_parser():
    eurosat_results = []
    url = 'https://eshop.eurosat.cz/kategorie/3697/cctv_kamery/'
    driver.get(url)
    time.sleep(2)
    button_hikvision = driver.find_element_by_xpath('//*[@id="supplier_26900"]')
    button_hikvision.click()
    time.sleep(2)
    button_select = driver.find_element_by_xpath('//*[@id="pavg_grid_id_toppager_center"]/table/tbody/tr/td[1]/select')
    button_select.click()
    time.sleep(3)
    button_100 = driver.find_element_by_xpath('//*[@id="pavg_grid_id_toppager_center"]/table/tbody/tr/td[1]/select/option[5]')
    button_100.click()
    time.sleep(2)

    button_next = driver.find_element_by_class_name('ui-icon-seek-next')
    i = 1
    while i <= 3:
        time.sleep(3)
        button_next.click()

        html = driver.page_source
        soup = BeautifulSoup(html, 'lxml')
        cams = soup.find_all('div', class_='grid-tile-main')
        for x in cams:
            try:
                try:
                    try:
                        availability = x.find('div', class_='on-stock').get_text(strip=True).lower()
                    except:
                        availability = x.find('div', class_='no-on-stock').get_text(strip=True).lower()
                    if 'není' in availability:
                        availability = 'Не на складе'
                    else:
                        availability = 'На складе'
                except:
                    availability = None

                if availability != None:
                    name = x.find('h3').find('a').get_text(strip=True).lower()
                    href = x.find('h3').find('a').get('href').lower()
                    eurosat_results.append({'name': name, 'availability': availability, 'href': 'https://eshop.eurosat.cz'+href})
            except:
                pass
        i += 1

    for i in eurosat_results:
        i['name'] = i['name'].replace(',', '.')
    return eurosat_results



def viakom_comparison(kenex_parser_results):
    url_1 = 'https://www.viakom.cz/kategorie/kamery-ip/#vendors=43,93&page=1&onstock=0&tabselect=0&navdatafilter=1,1,0&sortpar=16&sortdir=asc&ajx=true&navdata=1,1&view=table&rows=12'
    number_1 = 165
    url_2 = 'https://www.viakom.cz/kamery/turbo-hd-a-4v1-kamery/n-1,2,0#view=table&rows=12&page=1&onstock=0&tabselect=0&navdatafilter=1,2,0&sortpar=16&sortdir=asc&ajx=true&navdata=1,2,0&vendors=43,93,38'
    number_2 = 100
    viakom_parser_results = viakom_parser(url_1, number_1) + viakom_parser(url_2, number_2)
    viakom_comparison_list = []

    for k in kenex_parser_results:
        for v in viakom_parser_results:

            if 'hikvision ds-2cd2143g0-i (4mm)' in k['name']:
                if 'černá' not in k['name']:
                    k['name'] = 'hikvision ds-2cd2143g0-i *(4mm) venkovní 4 mp ip kamera'
            if 'hikvision ds-2cd2143g0-i (4mm)' == v['name']:
                v['name'] = 'hikvision ds-2cd2143g0-i *(4mm)'

            if 'hikvision ds-2cd2323g0-i (2.8mm)' in k['name']:
                if 'šedá' not in k['name']:
                    k['name'] = 'hikvision ds-2cd2323g0-i *(2.8mm) venkovní 2 mp ip kamera'
            if 'hikvision ds-2cd2323g0-i (2.8mm)' == v['name']:
                v['name'] = 'hikvision ds-2cd2323g0-i *(2.8mm)'

            kenex_replaced = k['name'].replace(' ', '').replace('.', '')
            viakom_replaced = v['name'].replace(' ', '').replace('.', '')

            if viakom_replaced in kenex_replaced:
                viakom_comparison_list.append({'name': k['name'], 'availability_kenex': k['availability'],\
                'availability_viakom': v['availability'],'href_kenex': k['href'], 'href_viakom': v['href']})

    i = 0
    while i <= len(viakom_comparison_list) - 1:
        if "camvia 4v1 cv-102" in viakom_comparison_list[i]['name']:
            viakom_comparison_list.pop(i)
            continue
        i += 1
    return viakom_comparison_list


def eurosat_comparison(kenex_parser_results):
    eurosat_parser_results = eurosat_parser()
    eurosat_comparison_list = []

    for k in kenex_parser_results:
        for e in eurosat_parser_results:

            if e['name'] == "ds-2cc12d9t-a" or e['name'] == "ds-2ce16h1t-it5" or e['name'] == "ds-2ce16h5t-it3z" or e['name'] == "ds-2ce56h5t-vpit3z":
                eurosat_parser_results.remove(e)
            if e['name'] == "ds-2ce56d8t-it3ze":
                e['name'] = "ds-2ce56d8t-it3ze 2.8"

            kenex_replaced = k['name'].replace(' ', '').replace('(', '').replace(')', '').replace('.', '').replace('-', '')
            eurosat_replaced = e['name'].replace(' ', '').replace('(outdoor)', '').replace('(', '').replace(')', '').replace('.', '').replace('-', '').replace('/', '').replace('_', '')

            if eurosat_replaced in kenex_replaced:
                eurosat_comparison_list.append({'name': k['name'], 'availability_kenex': k['availability'],\
                'availability_eurosat': e['availability'],'href_kenex': k['href'], 'href_eurosat': e['href']})
    return eurosat_comparison_list


def joining():
    kenex_parser_results = kenex_parser()
    viakom = viakom_comparison(kenex_parser_results)
    eurosat = eurosat_comparison(kenex_parser_results)
    joining_list = []

    for v in viakom:
        joining_list.append(v)
    for e in eurosat:
        joining_list.append(e)

    joining_list.sort(key = lambda n: n['name'])
    return joining_list


def cam():
    results = joining()
    sorted_results = []

    for i in results:
        if len(sorted_results) == 0:
            if 'href_viakom' in i.keys():
                try:
                    if i['availability_kenex'] != i['availability_viakom']:
                        result = {'name': i['name'], 'availability_kenex': i['availability_kenex'],\
                        'availability_eurosat': 'Нет в продаже', 'availability_viakom': i['availability_viakom'],\
                        'href_kenex': i['href_kenex'], 'href_eurosat': '—', 'href_viakom': i['href_viakom']}
                        sorted_results.append(result)
                except:
                    print('ОШИБКА 1')

            elif 'href_eurosat' in i.keys():
                try:
                    if i['availability_kenex'] != i['availability_eurosat']:
                        result = {'name': i['name'], 'availability_kenex': i['availability_kenex'],\
                        'availability_eurosat': i['availability_eurosat'], 'availability_viakom': 'Нет в продаже',\
                        'href_kenex': i['href_kenex'], 'href_eurosat': i['href_eurosat'], 'href_viakom': '—'}
                        sorted_results.append(result)
                except:
                    print('ОШИБКА 2')

        else:
            if i['name'] != sorted_results[-1]['name']:
                if 'href_viakom' in i.keys():
                    try:
                        if i['availability_kenex'] != i['availability_viakom']:
                            result = {'name': i['name'], 'availability_kenex': i['availability_kenex'],\
                            'availability_eurosat': 'Нет в продаже', 'availability_viakom': i['availability_viakom'],\
                            'href_kenex': i['href_kenex'], 'href_eurosat': '—', 'href_viakom': i['href_viakom']}
                            sorted_results.append(result)
                    except:
                        print('ОШИБКА 3')

                elif 'href_eurosat' in i.keys():
                    try:
                        if i['availability_kenex'] != i['availability_eurosat']:
                            result = {'name': i['name'], 'availability_kenex': i['availability_kenex'],\
                            'availability_eurosat': i['availability_eurosat'], 'availability_viakom': 'Нет в продаже',\
                            'href_kenex': i['href_kenex'], 'href_eurosat': i['href_eurosat'], 'href_viakom': '—'}
                            sorted_results.append(result)
                    except:
                        print('ОШИБКА 4')

            elif i['name'] == sorted_results[-1]['name']:
                try:
                    if i['availability_kenex']=='На складе' and i['availability_eurosat']=='Не на складе' and sorted_results[-1]['availability_viakom']=='Не на складе':
                        result = {'name': i['name'], 'availability_kenex': 'На складе',\
                        'availability_eurosat': 'Не на складе', 'availability_viakom': 'Не на складе',\
                        'href_kenex': i['href_kenex'], 'href_eurosat': i['href_eurosat'], 'href_viakom': sorted_results[-1]['href_viakom']}
                        sorted_results.append(result)
                        del sorted_results[-2]

                    elif i['availability_kenex']=='Не на складе' and i['availability_eurosat']=='На складе' and sorted_results[-1]['availability_viakom']=='На складе':
                        result = {'name': i['name'], 'availability_kenex': 'Не на складе',\
                        'availability_eurosat': 'На складе', 'availability_viakom': 'На складе',\
                        'href_kenex': i['href_kenex'], 'href_eurosat': i['href_eurosat'], 'href_viakom': sorted_results[-1]['href_viakom']}
                        sorted_results.append(result)
                        del sorted_results[-2]

                    elif i['availability_kenex']=='Не на складе' and i['availability_eurosat']=='Не на складе' and sorted_results[-1]['availability_viakom']=='На складе':
                        result = {'name': i['name'], 'availability_kenex': 'Не на складе',\
                        'availability_eurosat': 'Не на складе', 'availability_viakom': 'На складе',\
                        'href_kenex': i['href_kenex'], 'href_eurosat': i['href_eurosat'], 'href_viakom': sorted_results[-1]['href_viakom']}
                        sorted_results.append(result)
                        del sorted_results[-2]

                    elif i['availability_kenex']=='Не на складе' and i['availability_eurosat']=='На складе' and sorted_results[-1]['availability_viakom']=='Не на складе':
                        result = {'name': i['name'], 'availability_kenex': 'Не на складе',\
                        'availability_eurosat': 'На складе', 'availability_viakom': 'Не на складе',\
                        'href_kenex': i['href_kenex'], 'href_eurosat': i['href_eurosat'], 'href_viakom': sorted_results[-1]['href_viakom']}
                        sorted_results.append(result)
                        del sorted_results[-2]

                    else:
                        del sorted_results[-1]
                except:
                    print('ОШИБКА 5')

    if not sorted_results:
        text = '📹<b>Камеры:</b>\n\n'+'Доступность камер на Kenex соответствует доступности на Viakom и Eurosat✅'
        return text
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb['Sheet'])
        wb.create_sheet(title='Page {number}'.format(number=1))
        sheet = wb['Page {number}'.format(number=1)]
        cell_name = sheet.cell(row=1, column=1)
        cell_name.value = 'Название:'
        cell_availability_kenex = sheet.cell(row=1, column=2)
        cell_availability_kenex.value = 'Наличие Kenex:'
        cell_availability_eurosat = sheet.cell(row=1, column=3)
        cell_availability_eurosat.value = 'Наличие Eurosat:'
        cell_availability_viakom = sheet.cell(row=1, column=4)
        cell_availability_viakom.value = 'Наличие Viakom:'
        cell_href_kenex = sheet.cell(row=1, column=5)
        cell_href_kenex.value = 'Ссылка Kenex:'
        cell_href_eurosat = sheet.cell(row=1, column=6)
        cell_href_eurosat.value = 'Ссылка Eurosat:'
        cell_href_viakom = sheet.cell(row=1, column=7)
        cell_href_viakom.value = 'Ссылка Viakom:'

        for i in sorted_results:
            cell_1 = sheet.cell(row=sorted_results.index(i) + 3, column=1)
            cell_1.value = i['name']
            cell_2 = sheet.cell(row=sorted_results.index(i) + 3, column=2)
            cell_2.value = i['availability_kenex']
            cell_3 = sheet.cell(row=sorted_results.index(i) + 3, column=3)
            cell_3.value = i['availability_eurosat']
            cell_4 = sheet.cell(row=sorted_results.index(i) + 3, column=4)
            cell_4.value = i['availability_viakom']
            cell_5 = sheet.cell(row=sorted_results.index(i) + 3, column=5)
            cell_5.value = i['href_kenex']
            cell_6 = sheet.cell(row=sorted_results.index(i) + 3, column=6)
            cell_6.value = i['href_eurosat']
            cell_7 = sheet.cell(row=sorted_results.index(i) + 3, column=7)
            cell_7.value = i['href_viakom']

        wb.save('Камеры.xlsx')
